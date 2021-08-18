import sys
from copy import copy
from pathlib import Path

import openpyxl

file_path = sys.argv[1]
filter_col_name = sys.argv[2]
filter_file_location = sys.argv[3]

xlsx_file = Path('SimData', file_path)
src_workbook = openpyxl.load_workbook(xlsx_file)

all_sheets_name = src_workbook.sheetnames
print(f"Sheets Names : {all_sheets_name}")

filter_col_values = []
for src_worksheet_name in all_sheets_name:
    src_worksheet = src_workbook[src_worksheet_name]
    filter_col_index = 0
    for src_row in src_worksheet.iter_rows(max_row=1):
        for src_cell in src_row:
            if src_cell.value == filter_col_name:
                break
            filter_col_index = filter_col_index + 1

    print(f"Filter Column Index {filter_col_index} for sheet {src_worksheet_name}")

    current_col_index = 0
    for src_row in src_worksheet.iter_rows(min_row=None, max_row=None, min_col=None, max_col=None,
                                           values_only=False):
        for src_cell in src_row:
            if (src_cell.value != filter_col_name) & (current_col_index == filter_col_index):
                filter_col_values.append(src_cell.value)
            current_col_index = current_col_index + 1
        current_col_index = 0
filter_col_unique_values = set(filter_col_values)
print(f"Unique Filter Values : {filter_col_unique_values}")

for filter_col_unique_value in filter_col_unique_values:
    des_workbook = openpyxl.Workbook()
    des_workbook.remove(des_workbook["Sheet"])
    des_row_index = 1
    for src_worksheet_name in all_sheets_name:
        des_worksheet = des_workbook.create_sheet(src_worksheet_name)
        src_worksheet = src_workbook[src_worksheet_name]
        filter_col_index = 0
        for src_row in src_worksheet.iter_rows(max_row=1):
            for src_cell in src_row:
                if src_cell.value == filter_col_name:
                    break
                filter_col_index = filter_col_index + 1

        des_cell_index = 1
        for src_row in src_worksheet.iter_rows(min_row=None, max_row=None, min_col=None, max_col=None,
                                               values_only=False):
            if (src_row[filter_col_index].value == filter_col_unique_value) | (
                    src_row[filter_col_index].value == filter_col_name):
                for src_cell in src_row:
                    new_cell = des_worksheet.cell(row=des_row_index, column=des_cell_index, value=src_cell.value)
                    new_cell.data_type = src_cell.data_type
                    if src_cell.has_style:
                        new_cell.font = copy(src_cell.font)
                        new_cell.border = copy(src_cell.border)
                        new_cell.fill = copy(src_cell.fill)
                        new_cell.number_format = copy(src_cell.number_format)
                        new_cell.protection = copy(src_cell.protection)
                        new_cell.alignment = copy(src_cell.alignment)
                        new_cell._style = copy(src_cell._style)
                    des_cell_index = des_cell_index + 1
                des_cell_index = 1
                des_row_index = des_row_index + 1

        if des_row_index == 2:
            des_workbook.remove(des_worksheet)

        des_row_index = 1
    des_workbook.save(f'{filter_file_location}\Split-{filter_col_unique_value}.xlsx')
    des_workbook.close()
    print(f"Excel file 'Split-{filter_col_unique_value}.xlsx' created.")
print(f"Excel file data split into {filter_col_unique_values.__len__()} excel files successfully.")
